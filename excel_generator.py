import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from config import ClassConfig, CategoryConfig

def generate_grade_file(config: ClassConfig):
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{config.class_name} Grade Calculator"
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Add header
    ws['A1'] = f"{config.class_name} Grade Calculator"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add column headers
    ws['A2'] = "Category"
    ws['B2'] = "Percentage (%)"
    ws['D2'] = "Contribution (%)"
    
    for cell in ['A2', 'B2', 'D2']:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Create data validation for percentage inputs (0-100)
    dv = DataValidation(type="decimal", operator="between", formula1=0, formula2=100)
    dv.errorTitle = "Invalid Input"
    dv.error = "Please enter a value between 0 and 100"
    ws.add_data_validation(dv)
    
    # Highlight input cells
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # Function to add a category with multiple items
    def add_category(ws, start_row, category):
        title = f"{category.name} ({category.weight}%)"
        ws.merge_cells(f"A{start_row}:B{start_row}")
        ws[f"A{start_row}"] = title
        ws[f"A{start_row}"].font = Font(bold=True)
        
        for i in range(category.total_items):
            row = start_row + 1 + i
            ws[f"A{row}"] = f"Item {i+1}"
            ws[f"B{row}"] = 0  # Default percentage
            ws[f"B{row}"].fill = yellow_fill
            dv.add(f"B{row}")
        
        formula_row = start_row + category.total_items + 1
        ws[f"A{formula_row}"] = f"Best {category.best_of} Average:"
        ws[f"A{formula_row}"].font = Font(bold=True)
        
        if category.best_of < category.total_items:
            # Create a LARGE formula to get the best scores
            large_parts = []
            for j in range(1, category.best_of + 1):
                large_parts.append(str(j))
            large_formula = ",".join(large_parts)
            
            # Calculate average of top percentages
            ws[f"B{formula_row}"] = f"=AVERAGE(LARGE(B{start_row+1}:B{start_row+category.total_items},{{{large_formula}}}))"
        else:
            # If best_of equals total_items, just average all
            ws[f"B{formula_row}"] = f"=AVERAGE(B{start_row+1}:B{start_row+category.total_items})"
        
        # Calculate contribution to final grade
        ws[f"D{formula_row}"] = f"=B{formula_row}*{category.weight/100}"
        ws[f"D{formula_row}"].font = Font(bold=True)
        
        return formula_row + 2  # Return the next available row
    
    # Process all categories
    current_row = 3
    contribution_cells = []
    
    for category in config.categories:
        if category.total_items == 1:
            # Single item category (like Final Exam)
            ws[f"A{current_row}"] = f"{category.name} ({category.weight}%)"
            ws[f"A{current_row}"].font = Font(bold=True)
            ws[f"B{current_row}"] = 0  # Default percentage
            ws[f"B{current_row}"].fill = yellow_fill
            dv.add(f"B{current_row}")
            ws[f"D{current_row}"] = f"=B{current_row}*{category.weight/100}"
            contribution_cells.append(f"D{current_row}")
            current_row += 2
        else:
            # Multi-item category
            end_row = add_category(ws, current_row, category)
            contribution_cells.append(f"D{end_row-2}")  # The contribution cell is 2 rows above the returned row
            current_row = end_row
    
    # Add Final Grade calculation
    ws[f"A{current_row}"] = "Final Grade:"
    ws[f"A{current_row}"].font = Font(bold=True, size=12)
    ws[f"D{current_row}"] = f"=SUM({','.join(contribution_cells)})"
    ws[f"D{current_row}"].font = Font(bold=True, size=12)
    
    # Add formatting
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for row in ws.iter_rows(min_row=1, max_row=current_row, min_col=1, max_col=4):
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border
    
    # Save the workbook
    wb.save(f"{config.class_name}_Grade_Calculator.xlsx")
    print(f"Grade calculator created successfully as '{config.class_name}_Grade_Calculator.xlsx'")

def create_mat137_calculator():
    """Create the MAT137 grade calculator as a demonstration"""
    config = ClassConfig(
        class_name="MAT137",
        categories=[
            CategoryConfig("Pre-course Module", 2, 1, 1),
            CategoryConfig("Computation Quizzes", 3, 1, 1),
            CategoryConfig("Tutorial Worksheets", 5, 23, 17),
            CategoryConfig("Pre-class Quizzes", 5, 70, 50),
            CategoryConfig("Problem Sets", 12, 8, 6),
            CategoryConfig("Term Tests", 39, 4, 3),
            CategoryConfig("Final Exam", 34, 1, 1)
        ]
    )
    generate_grade_file(config)
